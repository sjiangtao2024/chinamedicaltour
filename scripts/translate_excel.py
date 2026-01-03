#!/usr/bin/env python3
import argparse
import csv
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell import MergedCell


CJK_RE = re.compile(r"[\u4e00-\u9fff]")


def has_cjk(value: str) -> bool:
    return bool(CJK_RE.search(value))


def export_translations(input_path: Path, output_path: Path) -> int:
    wb = load_workbook(input_path)
    rows = []

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.strip() and has_cjk(value):
                    rows.append(
                        {
                            "sheet": sheet.title,
                            "cell": cell.coordinate,
                            "original": value,
                            "translation": "",
                        }
                    )

    if not rows:
        print("No Chinese text found; nothing to export.")
        return 0

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f, fieldnames=["sheet", "cell", "original", "translation"]
        )
        writer.writeheader()
        writer.writerows(rows)

    print(f"Exported {len(rows)} rows to {output_path}")
    return 0


def load_translation_map(translations_path: Path) -> dict[tuple[str, str], str]:
    translations = {}
    with translations_path.open("r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        required = {"sheet", "cell", "original", "translation"}
        if not required.issubset(reader.fieldnames or []):
            missing = sorted(required - set(reader.fieldnames or []))
            raise ValueError(f"translations file missing columns: {missing}")
        for row in reader:
            key = (row["sheet"].strip(), row["cell"].strip())
            translation = row["translation"].strip()
            if translation:
                translations[key] = translation
    return translations


def apply_translations(
    input_path: Path, translations_path: Path, output_path: Path
) -> int:
    wb = load_workbook(input_path)
    translations = load_translation_map(translations_path)
    missing = 0

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.strip() and has_cjk(value):
                    key = (sheet.title, cell.coordinate)
                    translation = translations.get(key)
                    if translation:
                        cell.value = translation
                    else:
                        missing += 1

        # Column moves: K->M, L->N, O->Q, P->R
        for r in range(1, sheet.max_row + 1):
            copy_cell_value(sheet, f"K{r}", f"M{r}")
            copy_cell_value(sheet, f"L{r}", f"N{r}")
            copy_cell_value(sheet, f"O{r}", f"Q{r}")
            copy_cell_value(sheet, f"P{r}", f"R{r}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    if missing:
        print(
            f"Warning: {missing} Chinese cells had no translation; left unchanged."
        )

    # Quick sanity check: report any remaining Chinese text.
    remaining = 0
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.strip() and has_cjk(value):
                    remaining += 1

    if remaining:
        print(f"Remaining Chinese cells after apply: {remaining}")
    else:
        print("All Chinese text replaced.")

    print(f"Saved translated file to {output_path}")
    return 0


def get_merged_top_left(sheet, row: int, col: int) -> tuple[int, int] | None:
    for cell_range in sheet.merged_cells.ranges:
        if (
            cell_range.min_row <= row <= cell_range.max_row
            and cell_range.min_col <= col <= cell_range.max_col
        ):
            return (cell_range.min_row, cell_range.min_col)
    return None


def get_cell_value(sheet, cell_ref: str):
    cell = sheet[cell_ref]
    if isinstance(cell, MergedCell):
        top_left = get_merged_top_left(sheet, cell.row, cell.column)
        if top_left:
            return sheet.cell(row=top_left[0], column=top_left[1]).value
    return cell.value


def set_cell_value(sheet, cell_ref: str, value) -> None:
    cell = sheet[cell_ref]
    if isinstance(cell, MergedCell):
        top_left = get_merged_top_left(sheet, cell.row, cell.column)
        if top_left and (cell.row, cell.column) == top_left:
            sheet.cell(row=top_left[0], column=top_left[1]).value = value
        return
    cell.value = value


def copy_cell_value(sheet, src_ref: str, dst_ref: str) -> None:
    value = get_cell_value(sheet, src_ref)
    set_cell_value(sheet, dst_ref, value)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Export/apply translations for an Excel file."
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    export_parser = subparsers.add_parser("export", help="export translations")
    export_parser.add_argument("--input", required=True, type=Path)
    export_parser.add_argument("--output", required=True, type=Path)

    apply_parser = subparsers.add_parser("apply", help="apply translations")
    apply_parser.add_argument("--input", required=True, type=Path)
    apply_parser.add_argument("--translations", required=True, type=Path)
    apply_parser.add_argument("--output", required=True, type=Path)

    args = parser.parse_args()

    if args.command == "export":
        return export_translations(args.input, args.output)
    if args.command == "apply":
        return apply_translations(args.input, args.translations, args.output)

    return 1


if __name__ == "__main__":
    sys.exit(main())
